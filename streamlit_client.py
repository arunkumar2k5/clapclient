import streamlit as st
import csv
import io
import json
import asyncio
import os
import pandas as pd
import websockets
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from digikeyjson import fetch_digikey_data

SERVER_URL = os.getenv("MCP_SERVER_URL", "ws://127.0.0.1:8765")
SYSTEM_PROMPT = "You are an electronics expert. Compare the following components and provide a concise table format to state the parameters, highlighting key differences."
MODEL_NAME = "gpt-4o-mini"
TEMPERATURE = 0.2

def parse_llm_justifications(justification_text, all_params):
    """Parse LLM response and extract parameter-specific justifications with robust error handling."""
    justification_map = {}
    
    # Log the raw response for debugging
    st.write(f"**LLM Response Length:** {len(justification_text)} characters")
    
    # Try to parse markdown table
    lines = justification_text.split('\n')
    table_found = False
    
    for line in lines:
        # Check if this is a table row
        if '|' in line:
            # Skip separator lines (e.g., |---|---|)
            if line.strip().replace('|', '').replace('-', '').replace(' ', '').replace(':', '') == '':
                continue
            
            # Split by | and clean up
            parts = [p.strip() for p in line.split('|')]
            # Remove empty parts from start/end
            parts = [p for p in parts if p]
            
            # Skip header row
            if len(parts) >= 2 and parts[0].lower() in ['parameter', 'parameters']:
                table_found = True
                continue
            
            # Extract parameter and justification
            if len(parts) >= 2 and table_found:
                param_name = parts[0]
                justification = parts[1] if len(parts) > 1 else ""
                
                # Clean up the parameter name and justification
                param_name = param_name.strip()
                justification = justification.strip()
                
                if param_name and justification:
                    justification_map[param_name] = justification
    
    # If no table found, try alternative parsing (look for "Parameter:" patterns)
    if not justification_map:
        st.warning("⚠️ No markdown table found, trying alternative parsing...")
        current_param = None
        current_justification = []
        
        for line in lines:
            line = line.strip()
            # Look for parameter patterns like "**Parameter:**" or "Parameter:"
            if line.startswith('**') and '**' in line[2:]:
                if current_param and current_justification:
                    justification_map[current_param] = ' '.join(current_justification)
                current_param = line.split('**')[1].replace(':', '').strip()
                current_justification = []
            elif current_param and line:
                current_justification.append(line)
        
        # Add the last parameter
        if current_param and current_justification:
            justification_map[current_param] = ' '.join(current_justification)
    
    st.write(f"**Parsed {len(justification_map)} parameter justifications**")
    
    # Map justifications to parameters with fuzzy matching
    justifications = []
    unmatched_params = []
    
    for param in all_params:
        matched = False
        
        # Try exact match first
        if param in justification_map:
            justifications.append(justification_map[param])
            matched = True
        else:
            # Try case-insensitive match
            for key, value in justification_map.items():
                if key.lower() == param.lower():
                    justifications.append(value)
                    matched = True
                    break
            
            # Try partial match (key contains param or param contains key)
            if not matched:
                for key, value in justification_map.items():
                    if key.lower() in param.lower() or param.lower() in key.lower():
                        justifications.append(value)
                        matched = True
                        break
        
        if not matched:
            justifications.append("-")
            unmatched_params.append(param)
    
    if unmatched_params:
        st.warning(f"⚠️ Could not find justifications for: {', '.join(unmatched_params[:5])}")
        if len(unmatched_params) > 5:
            st.warning(f"   ... and {len(unmatched_params) - 5} more parameters")
    
    return justifications

def format_excel_sheet(worksheet, df):
    """Apply formatting to Excel worksheet."""
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for col_idx, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Format data rows
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            # Special formatting for Replacement Justification column (last column)
            if col_idx == len(df.columns):
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(df.columns, 1):
        column_letter = worksheet.cell(1, col_idx).column_letter
        
        if column == "Replacement Justification":
            worksheet.column_dimensions[column_letter].width = 80
        elif column == "Parameter":
            worksheet.column_dimensions[column_letter].width = 25
        else:
            # Calculate width based on content
            max_length = 0
            for row_idx in range(1, len(df) + 2):
                cell_value = str(worksheet.cell(row=row_idx, column=col_idx).value)
                max_length = max(max_length, len(cell_value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # Set row heights
    worksheet.row_dimensions[1].height = 30  # Header row
    for row_idx in range(2, len(df) + 2):
        worksheet.row_dimensions[row_idx].height = 30

async def send_json_to_llm(json_content: list, json_filename: str) -> dict:
    """Send JSON content to MCP server and get LLM analysis for replacement justification."""
    async with websockets.connect(SERVER_URL, max_size=2**23) as ws:
        await ws.send(json.dumps({
            "type": "initialize",
            "client": "streamlit-client",
            "version": "0.1"
        }))
        ready = json.loads(await ws.recv())
        if ready.get("type") != "ready":
            raise RuntimeError(f"Unexpected handshake response: {ready}")
        
        # Build prompt with reference component (first one)
        reference_part = json_content[0].get("Part Number", "Reference") if json_content else "Reference"
        prompt = f"""Analyze the following electronic components for replacement compatibility. 
The FIRST component '{reference_part}' is the REFERENCE component.

Components data:
{json.dumps(json_content, indent=2)}

Provide a markdown table with the following columns:
- Parameter: The parameter name
- Replacement Justification: A concise justification for each parameter

For each parameter, explain whether the alternative components are suitable replacements.
Focus on: electrical compatibility, mechanical compatibility, lifecycle status, and critical differences.
Be concise and specific. Format as a markdown table."""
        
        req = {
            "type": "request",
            "id": f"analyze-{json_filename}",
            "method": "llm.generate",
            "params": {
                "prompt": prompt,
                "system": SYSTEM_PROMPT,
                "model": MODEL_NAME,
                "temperature": TEMPERATURE,
                "format": "markdown"
            }
        }
        await ws.send(json.dumps(req))
        resp = json.loads(await ws.recv())
        if resp.get("type") == "result" and resp.get("ok"):
            return resp["data"]
        raise RuntimeError(f"Server error: {resp}")

st.set_page_config(page_title="CLAP", page_icon="🔗")
st.title("CLAP- Component Liketolike Assessment Platform")
st.write("Upload a CSV with Primary and Alternate part numbers")

uploaded_file = st.file_uploader("Upload CSV", type=["csv"])

if uploaded_file is not None:
    st.caption(f"Selected file: {uploaded_file.name}")
    if st.button("Process CSV"):
        try:
            text = uploaded_file.getvalue().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows_data = []
            for row in reader:
                part_numbers = []
                for col in ['Manf1_partno', 'Manf2_partno', 'Manf3_partno', 'Manf4_partno']:
                    pn = row.get(col, '').strip()
                    if pn:
                        part_numbers.append(pn)
                if part_numbers:
                    output_name = part_numbers[0]
                    rows_data.append((part_numbers, output_name))
        except Exception as exc:
            st.error(f"Unable to parse CSV: {exc}")
            rows_data = []
        if not rows_data:
            st.warning("No valid part numbers found in the uploaded CSV.")
        else:
            st.info(f"Found {len(rows_data)} row(s) to process. Fetching from Digikey...")
            json_files = []
            for part_numbers, output_name in rows_data:
                try:
                    filename = fetch_digikey_data(part_numbers, output_name)
                    json_files.append(filename)
                    st.success(f"Saved {filename} with {len(part_numbers)} part(s)")
                except Exception as exc:
                    st.error(f"{output_name}: {exc}")
            
            if json_files:
                st.info("Sending JSON files to LLM server for analysis...")
                excel_data = {}
                for fname in json_files:
                    try:
                        with open(fname, "r") as f:
                            json_content = json.load(f)
                        
                        # Create parameter comparison table
                        if not json_content:
                            st.warning(f"No data in {fname}")
                            continue
                        
                        # Get all unique parameters
                        all_params = set()
                        for component in json_content:
                            all_params.update(component.keys())
                        all_params = sorted(all_params)
                        
                        # Build dataframe with parameters as rows
                        data = {"Parameter": all_params}
                        for idx, component in enumerate(json_content):
                            part_num = component.get("Part Number", f"Component {idx+1}")
                            data[part_num] = [component.get(param, "-") for param in all_params]
                        
                        # Get LLM justification
                        st.info(f"🤖 Analyzing {fname} with LLM...")
                        llm_response = asyncio.run(send_json_to_llm(json_content, fname))
                        justification_text = llm_response.get("text", "No analysis available")
                        
                        # Show raw LLM response in an expander for debugging
                        with st.expander(f"📄 View Raw LLM Response for {fname}"):
                            st.text(justification_text)
                        
                        # Parse the LLM response with robust error handling
                        justifications = parse_llm_justifications(justification_text, all_params)
                        
                        data["Replacement Justification"] = justifications
                        
                        df = pd.DataFrame(data)
                        sheet_name = fname.replace(".json", "")[:31]
                        excel_data[sheet_name] = df
                        st.success(f"Analyzed {fname}")
                    except Exception as exc:
                        st.error(f"Error analyzing {fname}: {exc}")
                
                if excel_data:
                    output_excel = "analysis_results.xlsx"
                    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                        for sheet_name, df in excel_data.items():
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            # Apply formatting to the worksheet
                            worksheet = writer.sheets[sheet_name]
                            format_excel_sheet(worksheet, df)
                    st.success(f"Created Excel file: {output_excel}")
                    with open(output_excel, "rb") as f:
                        st.download_button(
                            label="Download Analysis Excel",
                            data=f,
                            file_name=output_excel,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                st.subheader("Download Individual JSON files:")
                for fname in json_files:
                    with open(fname, "rb") as f:
                        st.download_button(
                            label=f"Download {fname}",
                            data=f,
                            file_name=fname,
                            mime="application/json",
                            key=f"json_{fname}"
                        )
